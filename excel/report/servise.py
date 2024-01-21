import pandas as pd
from django.conf import settings
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pandas._typing import AggFuncType

from report.utils import funk_for_total_calk, funk_for_deviation, highlight


class ReportService:
    """
    Интерфейс для создания и форматирование корректного отчета на основе переданного excel файла
    """

    # Имя файла скорректированного отчета
    FILE_NAME = 'report'

    # Словарь для создания новых колонок в отчете
    NEW_COLUMNS = {
        'Исчислено всего по формуле': funk_for_total_calk,
        'Отклонения': funk_for_deviation
    }

    # Словарь для изменения названий колонок
    COLUMN_NAMES_DICT = {
        'Unnamed: 0': 'Филиал',
        'Unnamed: 1': 'Сотрудник',
        'Unnamed: 4': 'Налоговая база'
    }

    # Список для соединения колонок в шапке отчета
    MERGE_COLUMNS_INDEX = [
        (1, 2, 1, 1),
        (1, 2, 2, 2),
        (1, 2, 3, 3),
        (1, 1, 4, 5),
        (1, 2, 6, 6),
    ]

    # Список для изменения ширины колонок
    COLUMN_WIDTH = [
        ('A', 30),
        ('B', 20),
        ('C', 20),
        ('D', 20),
        ('E', 20),
        ('F', 20),
    ]
    # Список для изменения высоты строк
    ROW_HEIGHT = [
        {1: 12},
        {2: 27}
    ]
    # Список для форматирования шапки по координатам строки и ячейки
    HEADERS_RENGE = [2, 6]

    # Список колонок для изменения их оформления
    CELL_ALIGNMENT_LIST = ['A1', 'B1', 'C1', 'D1', 'D2', 'E2', 'F1']

    def __init__(self, file: str):
        self.file = file
        self.parser_file = ExcelParsers(file)

    def create_report(self):
        """
        Создание скорректированного отчета, его оформления и сохранение в media
        :return: путь к файлу
        """
        self.parser_file.rename_columns(self.COLUMN_NAMES_DICT)
        self.parser_file.del_column_by_value()
        self.create_columns()
        self.parser_file.sort_by_value()
        self.parser_file.create_style_to_df(highlight)
        filepath = self.parser_file.df_to_excel()
        filepath = self.formation_report(filepath)
        return filepath

    def create_columns(self) -> None:
        """
        Создает колонки используя NEW_COLUMNS, где:
        Ключ - название новой колонки
        Значение - функция, по которой будет создана колонка
        """
        for column_name, func in self.NEW_COLUMNS.items():
            self.parser_file.create_column_by_func(func, column_name)

    def formation_report(self, file_path: str) -> str:
        """
        Оформление скорректированного отчета
        :param file_path: путь к файлу который сохранил Excel Parsers
        :return: путь до оформленного отчета
        """
        ws = DesignReport(file_path)
        ws.merge_headers_cells(self.MERGE_COLUMNS_INDEX)
        ws.set_columns_name()
        ws.create_alignment_to_cells(self.CELL_ALIGNMENT_LIST)
        ws.set_cells_width(self.COLUMN_WIDTH)
        ws.set_row_height(self.ROW_HEIGHT)
        ws.format_all_cells(self.HEADERS_RENGE[1], size=10)
        ws.format_headers_cells(self.HEADERS_RENGE, size=10, bold=True)
        file_path = ws.save_document(file_path)

        return file_path


class ExcelParsers:
    """
    Класс для парсинга и создание скорректированного отчета
    """
    def __init__(self, report_file, engine: str = 'openpyxl'):
        """
        :param report_file: файл excel используя который будет создан скорректированный отчет
        :param engine: движок для парсинга файла
        """
        self.report_file = report_file
        self.engine = engine
        self.df = pd.read_excel(self.report_file, engine=self.engine, header=1, usecols=[0, 1, 4, 5])

    def rename_columns(self, colum_name_dict: dict) -> None:
        """
        Изменение названий колонок
        :param colum_name_dict: словарь, где:
            Ключ - имя колонки, которую нужно изменить
            Значение - новое имя для колонки
        """
        self.df.rename(columns=colum_name_dict, inplace=True)

    def del_column_by_value(self, value: str = 'Итого', column: str | int = 'Филиал', ) -> None:
        """
        Удаление из колонки строки по значению и обновляет self.df класса
        :param value: значение для поиска в колонке, если значение будет найдено в колонке - удаляет строку
        :param column: колонка в которой будет производиться поиск
        """
        column_index = self.df[self.df[column] == value].index
        self.df = self.df.drop(column_index, axis=0)

    def create_column_by_func(self, func: AggFuncType, column_name: str) -> None:
        """
        Создает новую колонку используя функцию и обновляет self.df класса
        :param func: функция с помощью которой будет создана новая колонка
        :param column_name: имя новой колонки
        """
        self.df[column_name] = self.df.apply(func, axis=1)

    def sort_by_value(self, value: str = 'Отклонения', ascending: bool = False) -> None:
        """
        Сортирует скорректированный отчет по заданной колонке и обновляет self.df класса
        :param value: колонка по которой будет отсортирован отчет
        :param ascending: направление сортировки:
            True - по возрастанию
            False - по убыванию
        """
        self.df = self.df.sort_values(by=[value], ascending=ascending)

    def create_style_to_df(self, func: AggFuncType, column_name: str = 'Отклонения', color: str = 'green') -> None:
        """
        Изменяет заливку ячеек в колонке используя функцию и обновляет self.df класса
        :param func: функция используя которую создается заливка ячейки
        :param column_name: колонка для заливки
        :param color: цвет заливки при положительно результате функции,
        при негативном результате цвет заливки - красный
        """
        self.df = self.df.style.applymap(func, color=color, subset=column_name)

    def df_to_excel(self, file_name: str = 'report', engine: str = 'openpyxl') -> str:
        """
        Сохранение скорректированного отчета в excel файл в медиа
        :param file_name: имя файла для сохранения
        :param engine: движок для excel файлов
        :return: путь к отчету в приложении 
        """
        filepath = settings.MEDIA_ROOT + file_name + '.xlsx'
        self.df.to_excel(filepath, engine=engine, index=False, sheet_name='Лист1')
        return filepath


class DesignReport:
    """
    Класс для оформления скорректированного отчета
    """
    def __init__(self, file_path: str):
        """
        :param file_path: путь до файла, который был сохранен ExcelParsers
        """
        self.file_path = file_path
        # чтение отчета
        self.wb = load_workbook(file_path)
        # переход на первый лист отчета
        self.ws = self.wb.active
        # получение списка заголовков отчета
        self.headers = [item.value for item in self.ws[1]]
        # добавление строки в начало файла 
        self.ws.insert_rows(idx=0, amount=1)

    @property
    def headers_dict(self) -> dict:
        """
        Формирование словаря с обновленными заголовками
        :return: словарь, где:
            Ключ - номер ячейки заголовка
            Значение - название ячейки
        """
        return {
            'A1': self.headers[0],
            'B1': self.headers[1],
            'C1': self.headers[2],
            'D1': 'Налог',
            'F1': self.headers[-1]
        }

    def set_columns_name(self) -> None:
        """
        Устанавливает название ячеек используя значения из headers_dict.
        Обновляет self.ws класса
        """

        for cell, value in self.headers_dict.items():
            self.ws[cell] = value

    def create_alignment_to_cells(self, cells_list: list[str]) -> None:
        """
        Изменяет оформление ячеек отчета и обновляет self.ws класса
        :param cells_list: список из координат колонок для изменения
        """
        _alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrapText=True
        )
        for cell in cells_list:
            self.ws[cell].alignment = _alignment

    def set_cells_width(self, colum_list: list[tuple[str, int]]) -> None:
        """
        Устанавливает ширину колонок и обновляет self.ws класса
        :param colum_list: лист с кортежами, где:
            tuple[0] - буквенная координата колонки
            tuple[1] - значение ширины колонки
        """
        for cell, _width in colum_list:
            self.ws.column_dimensions[cell].width = _width

    def set_row_height(self, list_rows: list[dict[int, int]]) -> None:
        """
        Устанавливает высоту строк в отчете и обновляет self.ws класса
        :param list_rows: лист со словарями, где:
            Ключ - номер строки
            Значение - высота строки
        """
        for item in list_rows:
            index = list(item.items())[0]
            self.ws.row_dimensions[index[0]].height = index[1]

    def merge_headers_cells(self, list_indexes: list[tuple[int, int, int, int]]) -> None:
        """
        Соединяет ячейки между собой и изменяет self.ws класса
        :param list_indexes: лист кортежей с координатами, где:
            tuple[0] - номер начальной строки
            tuple[1] - номер последней строки
            tuple[2] - номер начальной колонки
            tuple[3] - номер последней колонки
        """
        for indexes in list_indexes:
            self.ws.merge_cells(
                start_row=indexes[0],
                end_row=indexes[1],
                start_column=indexes[2],
                end_column=indexes[3]
            )

    def format_all_cells(self, index_end_cell: int, font: str = 'Arial', size: int = 11) -> None:
        """
        Форматирует шрифт всех колонок и обновляет self.ws класса
        :param index_end_cell: координаты последней колонки отчета
        :param font: стиль шрифта
        :param size: размер шрифта
        """
        _font = self.create_font(font, size)
        for cell in self.ws.iter_rows():
            for i in range(0, index_end_cell):
                cell[i].font = _font

    def format_headers_cells(
            self,
            list_indexes: list[int],
            font: str = 'Arial',
            size: int = 11,
            bold: bool = False,
            pattern: str = 'solid',
            fg_color: str = 'cbe4e5'
    ) -> None:
        """
        Изменение оформления заголовков и обновление self.ws класса
        :param list_indexes: лист с числовыми координатами максимальной строки и колонки заголовков
        :param font: стиль шрифта
        :param size: размер шрифта
        :param bold: Если True шрифт будет жирным
        :param pattern: шаблон для заливки ячейки
        :param fg_color: цвет заливки ячейки
        """
        _font = self.create_font(font, size, bold)
        _fill = self.create_fill(pattern, fg_color)
        for cell in self.ws.iter_rows(min_row=1, max_row=list_indexes[0], min_col=1, max_col=list_indexes[1]):
            for i in range(0, list_indexes[1]):
                cell[i].font = _font
                cell[i].fill = _fill

    @staticmethod
    def create_font(name: str, size: int, bold: bool = False) -> Font:
        """
        Создание объекта Fond для изменения стиля шрифта
        :param name: название стиля шрифта
        :param size: размер шрифта
        :param bold: Если True шрифт будет жирным
        :return: объект Font с заданными параметрами
        """
        return Font(
            name=name,
            size=size,
            bold=bold
        )

    @staticmethod
    def create_fill(pattern: str, fg_color: str) -> PatternFill:
        """
        Создание объекта PatternFill для заливки ячейки
        :param pattern: шаблон заливки
        :param fg_color: цвет заливки
        :return: объект PatternFill с заданными параметрами
        """
        return PatternFill(
            patternType=pattern,
            fgColor=fg_color
        )

    def save_document(self, filepath: str) -> str:
        """
        Сохраняет оформленный отчет в папку media c заменой скорректированного отчета сохраненного ExcelParsers
        :param filepath: путь по которому будет сохранен оформленный отчет
        :return: путь до файла
        """
        self.wb.save(filepath)
        self.wb.close()
        return filepath
